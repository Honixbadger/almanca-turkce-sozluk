"""
Excel'deki Almanca kelimeleri Türkçeye çevir.
1. Önce mevcut sözlükten çek
2. Bulunamayanlar için Google Translate (deep_translator)
"""
import json
import time
import random
import openpyxl
from deep_translator import GoogleTranslator

DICT_PATH = r"C:\Users\ozan\Desktop\almanca sözlük projesi\Playground-Yedek\almanca-sozluk-projesi\output\dictionary.json"
EXCEL_PATH = r"C:\Users\ozan\Desktop\almanca_quiz.xlsm"
CHECKPOINT_PATH = r"C:\Users\ozan\Desktop\almanca_sozluk_projesi\translate_checkpoint.json"

BATCH_SIZE = 30   # kelime/istek
DELAY_MIN = 0.8   # saniye
DELAY_MAX = 1.5


def load_dict_index():
    with open(DICT_PATH, encoding="utf-8") as f:
        d = json.load(f)
    idx = {}
    for entry in d:
        alm = entry.get("almanca", "").strip()
        art = entry.get("artikel", "").strip()
        tur = entry.get("turkce", "").strip()
        if not alm or not tur:
            continue
        idx[alm.lower()] = tur
        if art:
            idx[(art + " " + alm).lower()] = tur
    print(f"Sözlük indexi: {len(idx):,} kayıt")
    return idx


def load_checkpoint():
    try:
        with open(CHECKPOINT_PATH, encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return {}


def save_checkpoint(data):
    with open(CHECKPOINT_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def translate_batch(words, translator):
    """Kelimeleri | ile birleştirip tek istekte çevir."""
    joined = " | ".join(words)
    try:
        result = translator.translate(joined)
        parts = [p.strip() for p in result.split("|")]
        if len(parts) == len(words):
            return parts
        # Uzunluk uyuşmazlığı: tek tek dene
        return None
    except Exception as e:
        print(f"  Batch hata: {e}")
        return None


def translate_single(word, translator):
    try:
        result = translator.translate(word)
        return result
    except Exception as e:
        print(f"  Tek çeviri hata ({word}): {e}")
        return ""


def main():
    dict_idx = load_dict_index()
    checkpoint = load_checkpoint()

    # Excel kelimelerini oku
    wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)
    ws = wb["Kelimeler"]
    words = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        w = row[0]
        if w:
            words.append(w.strip())
    print(f"Excel'de {len(words):,} kelime")

    translator = GoogleTranslator(source="de", target="tr")

    # Çevirileri doldur
    translations = dict(checkpoint)

    # Sözlükten doldurulanlar
    dict_hits = 0
    for w in words:
        if w in translations:
            continue
        hit = dict_idx.get(w.lower())
        if hit:
            translations[w] = hit
            dict_hits += 1
    print(f"Sözlükten: {dict_hits:,} kelime | API gereken: {sum(1 for w in words if w not in translations):,}")

    # API ile çevrilecekler
    to_translate = [w for w in words if w not in translations]
    total = len(to_translate)

    if total == 0:
        print("Tüm kelimeler zaten çevrildi.")
    else:
        # Batch işleme
        batches = [to_translate[i:i+BATCH_SIZE] for i in range(0, total, BATCH_SIZE)]
        print(f"{len(batches)} batch ({BATCH_SIZE}'lik) işlenecek...")

        for bi, batch in enumerate(batches):
            results = translate_batch(batch, translator)

            if results:
                for word, tr in zip(batch, results):
                    translations[word] = tr
            else:
                # Tek tek dene
                for word in batch:
                    tr = translate_single(word, translator)
                    translations[word] = tr
                    time.sleep(0.3)

            # Her 10 batch'te checkpoint kaydet
            if (bi + 1) % 10 == 0:
                save_checkpoint(translations)
                done = (bi + 1) * BATCH_SIZE
                print(f"  {min(done, total)}/{total} çevrildi, checkpoint kaydedildi")

            time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

        save_checkpoint(translations)
        print(f"API çevirisi tamamlandı: {total:,} kelime")

    # Excel'e yaz
    print("Excel güncelleniyor...")
    wb2 = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)
    ws2 = wb2["Kelimeler"]

    for row in ws2.iter_rows(min_row=2):
        cell_de = row[0]
        cell_tr = row[1]
        if cell_de.value:
            word = cell_de.value.strip()
            tr = translations.get(word, "")
            cell_tr.value = tr

    wb2.save(EXCEL_PATH)
    print(f"Excel kaydedildi: {EXCEL_PATH}")
    print("TAMAMLANDI.")


if __name__ == "__main__":
    main()
