"""
PDF'den Almanca kelimeleri çıkar, lemmatize et, Excel'e yaz.
"""
import re
import sys
import fitz  # PyMuPDF
import spacy
import openpyxl
from collections import defaultdict

PDF_PATH = r"C:\Users\ozan\Desktop\Projeler\ozan\[ATZ_MTZ-Fachbuch ] Erich Hoepke, Stefan Breuer (eds.) - Nutzfahrzeugtechnik_ Grundlagen, Systeme, Komponenten (2016, Springer Vieweg) [10.1007_978-3-658-09537-6] - libgen.li.pdf"
EXCEL_PATH = r"C:\Users\ozan\Desktop\almanca_quiz.xlsm"

ARTICLE_MAP = {"Masc": "der", "Fem": "die", "Neut": "das"}

STOP_POS = {"PUNCT", "SPACE", "NUM", "X", "SYM"}

# Teknik/baskı gürültüsü için regex
JUNK_RE = re.compile(r'^[^a-zA-ZäöüÄÖÜß]|[^a-zA-ZäöüÄÖÜß]$')
VALID_WORD_RE = re.compile(r'^[a-zA-ZäöüÄÖÜß\-]{2,}$')


def extract_text_from_pdf(path):
    print("PDF okunuyor...")
    doc = fitz.open(path)
    pages = []
    for i, page in enumerate(doc):
        text = page.get_text("text")
        pages.append(text)
        if (i + 1) % 50 == 0:
            print(f"  {i+1}/{len(doc)} sayfa okundu")
    page_count = len(doc)
    doc.close()
    full_text = "\n".join(pages)
    print(f"Toplam {page_count} sayfa, {len(full_text):,} karakter çıkarıldı.")
    return full_text


def lemmatize_words(text):
    print("spaCy modeli yükleniyor (de_core_news_lg)...")
    nlp = spacy.load("de_core_news_lg")
    # Uzun metin için batch processing
    nlp.max_length = 10_000_000

    print("Lemmatizasyon başlıyor (bu birkaç dakika sürebilir)...")
    # Büyük metni parçalara böl
    chunk_size = 500_000
    chunks = [text[i:i+chunk_size] for i in range(0, len(text), chunk_size)]

    # lemma → (pos, gender) dict
    word_data = {}  # lemma_display -> pos

    for ci, chunk in enumerate(chunks):
        print(f"  Parça {ci+1}/{len(chunks)} işleniyor...")
        doc = nlp(chunk)
        for token in doc:
            if token.pos_ in STOP_POS:
                continue
            lemma = token.lemma_.strip()
            if not VALID_WORD_RE.match(lemma):
                continue
            if len(lemma) < 2:
                continue

            pos = token.pos_

            # İsimler için artikel ekle
            if pos == "NOUN":
                gender = token.morph.get("Gender")
                if gender:
                    artikel = ARTICLE_MAP.get(gender[0], "")
                    display = f"{artikel} {lemma}" if artikel else lemma
                else:
                    display = lemma
            elif pos == "VERB":
                display = lemma  # mastar formu zaten lemma
            elif pos == "ADJ":
                display = lemma  # temel form
            else:
                display = lemma

            if display not in word_data:
                word_data[display] = pos

    print(f"Toplam {len(word_data):,} benzersiz kelime bulundu.")
    return word_data


def write_to_excel(word_data, excel_path):
    print("Excel dosyası güncelleniyor...")
    wb = openpyxl.load_workbook(excel_path, keep_vba=True)
    ws = wb["Kelimeler"]

    # POS sırasına göre sırala: NOUN, VERB, ADJ, diğerleri, alfabetik
    POS_ORDER = {"NOUN": 0, "VERB": 1, "ADJ": 2, "ADV": 3}
    sorted_words = sorted(
        word_data.items(),
        key=lambda x: (POS_ORDER.get(x[1], 9), x[0].lower())
    )

    # Header satırı korunuyor (satır 1), 2. satırdan itibaren yaz
    for i, (word, pos) in enumerate(sorted_words):
        row = i + 2  # row 1 = header
        ws.cell(row=row, column=1, value=word)

    # Eğer mevcut veri daha uzunsa, fazlasını temizle
    last_data_row = len(sorted_words) + 1
    for row in range(last_data_row + 1, ws.max_row + 1):
        ws.cell(row=row, column=1, value=None)

    wb.save(excel_path)
    print(f"{len(sorted_words):,} kelime '{excel_path}' dosyasına yazıldı.")


if __name__ == "__main__":
    text = extract_text_from_pdf(PDF_PATH)
    word_data = lemmatize_words(text)
    write_to_excel(word_data, EXCEL_PATH)
    print("TAMAMLANDI.")
